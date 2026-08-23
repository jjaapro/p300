#!/usr/bin/env python3
"""Build the timeline workbook: Summary / Positions / Events / Signals / Scoreboard."""
import json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

POS = json.load(open('/tmp/paladin/extract/positions_all.json', encoding='utf-8'))
EVS = json.load(open('/tmp/paladin/extract/events_all.json', encoding='utf-8'))
PB = json.load(open('/tmp/paladin/extract/playbook_final.json', encoding='utf-8'))
NS = json.load(open('/tmp/paladin/extract/nosymbol_resolved.json', encoding='utf-8'))
OUT = '/tmp/paladin/out/paladin_trade_timeline.xlsx'

msgs = sorted(json.load(open('/tmp/paladin/scalpx_paladin.json', encoding='utf-8'))['messages'],
              key=lambda m: m['timestamp'])
LINK = {i + 1: f"https://discord.com/channels/1042855255089623100/1501271173588193290/{m['id']}"
        for i, m in enumerate(msgs)}

F = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='23303F')
HDR_FONT = Font(name=F, bold=True, color='FFFFFF', size=10)
BODY = Font(name=F, size=10)
THIN = Side(style='thin', color='DDDDDD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')


def clip(v, n=32000):
    if v is None or isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    return s if len(s) <= n else s[:n - 3] + '...'


def sheet(ws, headers, rows, widths, wrap_cols=(), table=None, heights=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32
    for r in rows:
        ws.append([clip(v) for v in r])
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = WRAP if cell.column in wrap_cols else TOP
        if heights:
            ws.row_dimensions[row[0].row].height = heights
    ws.freeze_panes = 'A2'
    if table and ws.max_row > 1:
        t = Table(displayName=table, ref=f'A1:{get_column_letter(len(headers))}{ws.max_row}')
        t.tableStyleInfo = TableStyleInfo(name='TableStyleLight11', showRowStripes=True)
        ws.add_table(t)


wb = Workbook()

# ------------------------------------------------------------- Positions ----
ws = wb.active
ws.title = 'Positions'
H = ['#', 'First seen', 'Symbol', 'Side', 'Outcome', 'Exit type', 'R', 'Card ROI %', 'Leverage',
     'Planned entry', 'Entry note', 'Planned DCA', 'Planned stop', 'Planned targets', 'Risk note',
     'Avg entry', 'DCA filled', 'Stop moves', 'Target changes', 'Partials taken', 'Exit price',
     'Entry time', 'Exit time', 'Held (h)', 'Thesis', 'What he did', 'Indicators', 'Market context',
     'Still open', 'Confidence', 'Trust notes', 'Messages', 'First message link']
rows = []
for i, p in enumerate(sorted(POS, key=lambda x: x['first_seen']), 1):
    ms = p.get('msgs') or []
    rows.append([
        i, p['first_seen'], p['symbol'], p['side'], p['outcome'], p.get('exit_type'),
        p.get('r_multiple'), p.get('roi_pct'), p.get('leverage'),
        p.get('planned_entry'), p.get('planned_entry_note'),
        ', '.join(str(x) for x in (p.get('planned_dca') or [])) or None,
        p.get('planned_stop'),
        ', '.join(str(x) for x in (p.get('planned_take_profits') or [])) or None,
        p.get('risk_note'), p.get('avg_entry'),
        ' | '.join(f"{f.get('price')}" + (f" ({f['note']})" if f.get('note') else '')
                   for f in (p.get('dca_fills') or [])) or None,
        ' -> '.join(f"{s.get('to')}" + (f" ({s['type']})" if s.get('type') else '')
                    for s in (p.get('stop_moves') or [])) or None,
        ' -> '.join('/'.join(str(x) for x in (c.get('to') or [])) for c in (p.get('tp_changes') or [])) or None,
        ' | '.join((f"{x['portion_pct']}% at " if x.get('portion_pct') else '') + str(x.get('price'))
                   for x in (p.get('partials') or [])) or None,
        p.get('exit_price'), p.get('entry_time'), p.get('exit_time'), p.get('duration_hours'),
        p.get('thesis'), p.get('management_story'), ', '.join(p.get('indicators') or []) or None,
        p.get('market_context'), 'yes' if p.get('still_open_at_end') else '', p.get('confidence'),
        p.get('notes'), ', '.join(f'#{m}' for m in ms), LINK.get(ms[0]) if ms else None,
    ])
sheet(ws, H, rows,
      widths=[5, 19, 11, 7, 11, 13, 7, 10, 9, 13, 26, 16, 12, 20, 24, 12, 26, 24, 20, 26, 12,
              19, 19, 9, 52, 74, 22, 26, 9, 10, 60, 30, 36],
      wrap_cols=(11, 15, 25, 26, 27, 28, 31), table='Positions', heights=44)

# ---------------------------------------------------------------- Events ----
ws2 = wb.create_sheet('Events')
H2 = ['Msg', 'Timestamp', 'Event', 'Symbol', 'Side', 'Entry', 'DCA', 'Targets', 'Stop',
      'Portion %', 'Leverage', 'R', 'ROI %', 'Price referenced', 'Risk note', 'Why he did it',
      'Indicators', 'Market context', 'What he wrote', 'Confidence', 'Link']
r2 = [[e['msg'], e['timestamp'], e['event_type'], e.get('symbol'), e.get('side'),
       e.get('entry_price'), e.get('dca_price'),
       ', '.join(str(x) for x in (e.get('take_profits') or [])) or None,
       e.get('stop_loss'), e.get('portion_pct'), e.get('leverage'), e.get('r_multiple'),
       e.get('roi_pct'), e.get('price_ref'), e.get('risk_note'), e.get('rationale'),
       ', '.join(e.get('indicators_mentioned') or []) or None, e.get('market_context'),
       e.get('quote'), e.get('confidence'), LINK.get(e['msg'])] for e in EVS]
sheet(ws2, H2, r2,
      widths=[6, 19, 14, 11, 7, 12, 12, 18, 12, 9, 9, 7, 9, 14, 22, 52, 22, 24, 62, 10, 36],
      wrap_cols=(15, 16, 17, 18, 19), table='Events', heights=32)

# --------------------------------------------------------------- Signals ----
ws3 = wb.create_sheet('What he watches')
H3 = ['Signal', 'Category', 'Messages', 'First seen', 'Last seen', 'What it is',
      'How he uses it', 'Evidence (his words)']


def ev_txt(item):
    return '\n\n'.join(f"#{e['msg']} ({e.get('date','')}): \"{e['quote']}\"" for e in (item.get('evidence') or []))


r3 = [[s['name'], s.get('category'), s.get('frequency'), s.get('first_seen'), s.get('last_seen'),
       s.get('what_it_is'), s.get('how_he_uses_it'), ev_txt(s)] for s in PB['signals']]
for group, key, det in [('Risk rule', 'rule', 'detail'), ('Entry pattern', 'pattern', 'detail'),
                        ('Management tactic', 'tactic', 'detail')]:
    src = {'Risk rule': PB['risk_rules'], 'Entry pattern': PB['entry_style'],
           'Management tactic': PB['trade_management']}[group]
    for it in src:
        r3.append([it[key], group, it.get('frequency'), None, None, None, it.get(det), ev_txt(it)])
sheet(ws3, H3, r3, widths=[46, 17, 10, 12, 12, 54, 66, 78],
      wrap_cols=(1, 6, 7, 8), table='Signals', heights=90)

# ------------------------------------------------------------ Scoreboard ----
ws4 = wb.create_sheet('His scoreboard')
H4 = ['Msg', 'Date', 'Period claimed', 'Claim', 'Wins', 'Losses', 'BE', 'Total', 'R or %', 'His words', 'Link']
r4 = [[c.get('msg'), c.get('date'), c.get('period_claimed'), c.get('claim'), c.get('wins'),
       c.get('losses'), c.get('breakeven'), c.get('total'), c.get('r_or_pct'), c.get('quote'),
       LINK.get(c.get('msg'))] for c in NS.get('self_reported_scoreboard', [])]
sheet(ws4, H4, r4, widths=[6, 12, 26, 52, 7, 8, 6, 7, 22, 66, 36],
      wrap_cols=(3, 4, 10), table='Scoreboard', heights=42)

# --------------------------------------------------------------- Summary ----
ws0 = wb.create_sheet('Summary', 0)
ws0.sheet_view.showGridLines = False
ws0.cell(row=1, column=1, value="Paladin (ScalpX) — reconstructed trade timeline").font = Font(name=F, bold=True, size=15)
ws0.cell(row=2, column=1, value="Built from all 1,005 messages in #paladin (2026-05-06 to 2026-08-21) plus the 320 screenshots he posted. "
                                "Blank means he never published the number — nothing here is calculated on his behalf.").font = Font(name=F, italic=True, size=10)

lab = Font(name=F, bold=True, size=10)
row = 4
ws0.cell(row=row, column=1, value='The record').font = Font(name=F, bold=True, size=12)
row += 1
first_stat_row = row
for label, formula in [
    ('Positions catalogued', '=COUNTA(Positions!C2:C100000)'),
    ('…actually taken by him', '=COUNTA(Positions!C2:C100000)-COUNTIF(Positions!E2:E100000,"not_taken")-COUNTIF(Positions!E2:E100000,"never_filled")'),
    ('…with a knowable outcome', '=COUNTIF(Positions!E2:E100000,"win")+COUNTIF(Positions!E2:E100000,"loss")+COUNTIF(Positions!E2:E100000,"breakeven")'),
    ('Wins', '=COUNTIF(Positions!E2:E100000,"win")'),
    ('Losses', '=COUNTIF(Positions!E2:E100000,"loss")'),
    ('Breakeven', '=COUNTIF(Positions!E2:E100000,"breakeven")'),
    ('Never resolved in the channel', '=COUNTIF(Positions!E2:E100000,"unknown")'),
    ('Called but not taken', '=COUNTIF(Positions!E2:E100000,"not_taken")+COUNTIF(Positions!E2:E100000,"never_filled")'),
]:
    ws0.cell(row=row, column=1, value=label).font = lab
    ws0.cell(row=row, column=2, value=formula).font = BODY
    row += 1
wr_row = row
ws0.cell(row=row, column=1, value='Win rate excluding breakeven').font = lab
c = ws0.cell(row=row, column=2, value=f'=B{first_stat_row+3}/(B{first_stat_row+3}+B{first_stat_row+4})')
c.font = BODY; c.number_format = '0.0%'
row += 1
ws0.cell(row=row, column=1, value='Win rate including breakeven').font = lab
c = ws0.cell(row=row, column=2, value=f'=B{first_stat_row+3}/B{first_stat_row+2}')
c.font = BODY; c.number_format = '0.0%'
row += 1
ws0.cell(row=row, column=1, value='He publicly claimed 85% (May) and 89% (July).').font = Font(name=F, italic=True, size=9, color='808080')

row += 2
ws0.cell(row=row, column=1, value='How trades ended').font = Font(name=F, bold=True, size=12)
row += 1
for label, formula, note in [
    ('Closed by hand', '=COUNTIF(Positions!F2:F100000,"manual_close")', 'of these, wins:'),
    ('Reached a published target', '=COUNTIF(Positions!F2:F100000,"take_profit")', ''),
    ('Hit a stop', '=COUNTIF(Positions!F2:F100000,"stop_loss")', 'every loss came this way'),
    ('Closed flat', '=COUNTIF(Positions!F2:F100000,"breakeven")', ''),
    ('Simply stopped being mentioned', '=COUNTIF(Positions!F2:F100000,"unknown")', ''),
]:
    ws0.cell(row=row, column=1, value=label).font = lab
    ws0.cell(row=row, column=2, value=formula).font = BODY
    if label == 'Closed by hand':
        ws0.cell(row=row, column=3, value='=COUNTIFS(Positions!F2:F100000,"manual_close",Positions!E2:E100000,"win")').font = BODY
        ws0.cell(row=row, column=4, value='wins, 0 losses').font = Font(name=F, size=9, color='808080')
    elif note:
        ws0.cell(row=row, column=3, value=note).font = Font(name=F, size=9, color='808080')
    row += 1

row += 1
ws0.cell(row=row, column=1, value='Did he publish a plan?').font = Font(name=F, bold=True, size=12)
row += 1
for label, formula in [
    ('Positions with a published stop', '=COUNT(Positions!M2:M100000)'),
    ('Positions with a published target', '=COUNTA(Positions!N2:N100000)'),
    ('Positions with a published DCA level', '=COUNTA(Positions!L2:L100000)'),
    ('…where that DCA actually filled', '=COUNTA(Positions!Q2:Q100000)'),
    ('Positions where he moved the stop', '=COUNTA(Positions!R2:R100000)'),
    ('Positions where he took partials', '=COUNTA(Positions!T2:T100000)'),
]:
    ws0.cell(row=row, column=1, value=label).font = lab
    ws0.cell(row=row, column=2, value=formula).font = BODY
    row += 1

row += 1
ws0.cell(row=row, column=1, value='Month by month').font = Font(name=F, bold=True, size=12)
row += 1
for i, h in enumerate(['Month', 'Wins', 'Losses', 'BE', 'Win rate excl. BE', 'What he claimed'], 1):
    ws0.cell(row=row, column=i, value=h).font = lab
row += 1
claims = {'2026-05': '60 trades, 51W, 9L, 85%, +159.6%', '2026-06': 'no summary ever posted',
          '2026-07': '45 trades, 40W, 5 SLs, 89%, +23.5R', '2026-08': 'log ends 21 Aug'}
for m in ['2026-05', '2026-06', '2026-07', '2026-08']:
    ws0.cell(row=row, column=1, value=m).font = BODY
    ws0.cell(row=row, column=2, value=f'=COUNTIFS(Positions!$B$2:$B$100000,"{m}*",Positions!$E$2:$E$100000,"win")').font = BODY
    ws0.cell(row=row, column=3, value=f'=COUNTIFS(Positions!$B$2:$B$100000,"{m}*",Positions!$E$2:$E$100000,"loss")').font = BODY
    ws0.cell(row=row, column=4, value=f'=COUNTIFS(Positions!$B$2:$B$100000,"{m}*",Positions!$E$2:$E$100000,"breakeven")').font = BODY
    c = ws0.cell(row=row, column=5, value=f'=IFERROR(B{row}/(B{row}+C{row}),"-")')
    c.font = BODY; c.number_format = '0%'
    ws0.cell(row=row, column=6, value=claims[m]).font = Font(name=F, size=9, color='808080')
    row += 1

row += 2
ws0.cell(row=row, column=1, value='The sheets').font = Font(name=F, bold=True, size=12)
row += 1
for line in [
    'Positions — one row per trade, from the call to the exit. The "planned" columns are what he published up front; the "avg entry / partials / exit" columns are what the messages and screenshots show he actually did.',
    'Events — all 1,242 individual actions (calls, entries, DCAs, stop moves, partials, exits) with his exact words and a link to the message.',
    'What he watches — the signals, risk rules, entry patterns and management tactics the record supports, each with his own quotes.',
    'His scoreboard — every performance claim he made, so his numbers can be checked against the Positions sheet.',
    'A "medium" or "low" confidence on a position means part of it was inferred; the Trust notes column says which part.',
]:
    ws0.cell(row=row, column=1, value=line).font = Font(name=F, size=10)
    ws0.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws0.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws0.row_dimensions[row].height = 30
    row += 1

for col, w in zip('ABCDEF', [40, 14, 14, 16, 18, 34]):
    ws0.column_dimensions[col].width = w

wb.save(OUT)
print('saved', OUT, '| positions', len(POS), '| events', len(EVS))
