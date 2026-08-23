#!/usr/bin/env python3
"""Build the paladin image-analysis workbook."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC = '/tmp/paladin/extract/merged.json'
OUT = '/tmp/paladin/out/paladin_image_analysis.xlsx'
merged = json.load(open(SRC, encoding='utf-8'))

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LONG = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')


def clip(v, n=32000):
    """Keep numbers numeric; only strings get stringified/truncated."""
    if v is None or isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    return s if len(s) <= n else s[:n - 3] + '...'


def write_sheet(ws, headers, rows, widths, wrap_cols=(), freeze='A2', table_name=None):
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    for r in rows:
        ws.append([clip(v) for v in r])
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = BODY
            cell.border = BORDER
            cell.alignment = LONG if cell.column in wrap_cols else TOP
    ws.freeze_panes = freeze
    if table_name and ws.max_row > 1:
        ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showRowStripes=True)
        ws.add_table(t)
    return ws


wb = Workbook()

# ---------------------------------------------------------------- Trades ----
ws = wb.active
ws.title = 'Trades'
headers = ['Date', 'Time', 'Symbol', 'Base', 'Side', 'Leverage', 'Margin mode', 'Entry price',
           'Exit price', 'Exit price type', 'ROI %', 'PnL USD', 'Position size', 'Margin USD',
           'Stop loss', 'Take profits', 'Liq. price', 'Status', 'Platform', 'Image type',
           'Trade notes', 'Message posted with image', 'Image file', 'Confidence', 'Discord link']
rows = []
for m in merged:
    for t in (m['trades'] or []):
        tp = t.get('take_profits') or []
        rows.append([
            m['date'], m['time'], t.get('symbol_norm') or t.get('symbol'), t.get('base'),
            t.get('side'), t.get('leverage'), t.get('margin_mode'), t.get('entry_price'),
            t.get('exit_price'), t.get('exit_price_type'), t.get('roi_pct'), t.get('pnl_usd'),
            t.get('position_size'), t.get('margin_usd'), t.get('stop_loss'),
            '; '.join(str(x) for x in tp) if tp else None, t.get('liquidation_price'),
            t.get('status'), m['platform'], m['image_type'], t.get('notes'), m['message'],
            m['filename'], m['confidence'], m['message_link'],
        ])
rows.sort(key=lambda r: (r[0], r[1]))
write_sheet(ws, headers, rows,
            widths=[11, 9, 12, 8, 7, 9, 11, 13, 13, 13, 9, 10, 16, 11, 11, 18, 11, 9, 10, 20, 46, 46, 40, 10, 34],
            wrap_cols=(21, 22, 23), table_name='TradesTbl')
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for i in (8, 9, 15, 17):
        row[i - 1].number_format = '#,##0.000000'
    row[10].number_format = '#,##0.00'
    row[11].number_format = '#,##0.00'
n_trades = ws.max_row - 1

# ----------------------------------------------------------------- Images ---
ws2 = wb.create_sheet('Images')
h2 = ['#', 'Date', 'Time', 'Trading related', 'Image type', 'Platform', 'Trades in image',
      'Symbols', 'Chart timeframe', 'Chart indicators', 'Chart levels / drawings',
      'What the chart shows', 'Description', 'Text visible in image',
      'Message posted with image', 'Date/time printed on image', 'Confidence', 'Issues',
      'Image file', 'Duplicate of', 'Discord link']
r2 = []
for m in merged:
    ch = m.get('chart') or {}
    syms = sorted({(t.get('symbol_norm') or t.get('symbol') or '') for t in (m['trades'] or [])} - {''})
    r2.append([m['index'], m['date'], m['time'], 'yes' if m['trading_related'] else 'no',
               m['image_type'], m['platform'], len(m['trades'] or []), ', '.join(syms) or None,
               ch.get('timeframe'), '; '.join(ch.get('indicators') or []) or None,
               '; '.join(ch.get('drawings_and_levels') or []) or None, ch.get('what_it_shows'),
               m['description'], m['visible_text'], m['message'], m['timestamp_on_image'],
               m['confidence'], m['issues'], m['filename'], m['duplicate_of'] or None,
               m['message_link']])
write_sheet(ws2, h2, r2,
            widths=[6, 11, 9, 9, 22, 10, 9, 18, 10, 30, 40, 46, 52, 60, 46, 22, 10, 34, 40, 34, 34],
            wrap_cols=(10, 11, 12, 13, 14, 15, 18), table_name='ImagesTbl')
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    ws2.row_dimensions[row[0].row].height = 42

# ----------------------------------------------------------------- Charts ---
ws3 = wb.create_sheet('Charts')
h3 = ['Date', 'Time', 'Symbol', 'Timeframe', 'Platform', 'Price shown', 'Indicators',
      'Levels / drawings', 'Annotations on chart', 'What it shows',
      'Message posted with image', 'Image file', 'Discord link']
r3 = []
for m in merged:
    ch = m.get('chart')
    if not ch:
        continue
    r3.append([m['date'], m['time'], ch.get('symbol'), ch.get('timeframe'), m['platform'],
               ch.get('price_shown'), '; '.join(ch.get('indicators') or []) or None,
               '; '.join(ch.get('drawings_and_levels') or []) or None,
               ch.get('annotations_text'), ch.get('what_it_shows'), m['message'],
               m['filename'], m['message_link']])
write_sheet(ws3, h3, r3, widths=[11, 9, 12, 10, 11, 14, 34, 46, 40, 60, 40, 40, 34],
            wrap_cols=(7, 8, 9, 10, 11), table_name='ChartsTbl')
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    ws3.row_dimensions[row[0].row].height = 60

# ---------------------------------------------------------------- Summary ---
ws4 = wb.create_sheet('Summary', 0)
title = ws4.cell(row=1, column=1, value='Paladin (ScalpX) — trading information extracted from posted images')
title.font = Font(name=FONT, bold=True, size=14)
ws4.cell(row=2, column=1, value='Channel #🔮｜paladin, 2026-05-06 to 2026-08-21. Every one of the 320 images (incl. 2 screen recordings) was viewed and transcribed; numbers are copied from the image, never inferred.').font = Font(name=FONT, size=10, italic=True)

lab = Font(name=FONT, bold=True, size=10)
row = 4
ws4.cell(row=row, column=1, value='Coverage').font = Font(name=FONT, bold=True, size=12)
row += 1
trade_rows_cell = None
for label, formula in [
    ('Images analysed', '=COUNTA(Images!A2:A100000)'),
    ('Images with trading information', '=COUNTIF(Images!D2:D100000,"yes")'),
    ('Images with no trading information', '=COUNTIF(Images!D2:D100000,"no")'),
    ('Trade rows extracted', '=COUNTA(Trades!C2:C100000)'),
    ('Price charts captured', '=COUNTA(Charts!A2:A100000)'),
    ('Rows flagged medium confidence', '=COUNTIF(Images!Q2:Q100000,"medium")'),
]:
    ws4.cell(row=row, column=1, value=label).font = lab
    c = ws4.cell(row=row, column=2, value=formula)
    c.font = BODY
    if label == 'Trade rows extracted':
        trade_rows_cell = f'$B${row}'
    row += 1

row += 1
ws4.cell(row=row, column=1, value='Field fill rate (of trade rows)').font = Font(name=FONT, bold=True, size=12)
row += 1
for label, formula in [
    ('Entry price', '=COUNT(Trades!H2:H100000)'),
    ('Exit / close price', '=COUNT(Trades!I2:I100000)'),
    ('ROI %', '=COUNT(Trades!K2:K100000)'),
    ('Leverage', '=COUNT(Trades!F2:F100000)'),
    ('Side (long/short)', '=COUNTIF(Trades!E2:E100000,"long")+COUNTIF(Trades!E2:E100000,"short")'),
    ('Stop loss', '=COUNT(Trades!O2:O100000)'),
    ('Take profit', '=COUNTA(Trades!P2:P100000)'),
    ('Liquidation price', '=COUNT(Trades!Q2:Q100000)'),
    ('PnL in USD', '=COUNT(Trades!L2:L100000)'),
]:
    ws4.cell(row=row, column=1, value=label).font = lab
    ws4.cell(row=row, column=2, value=formula).font = BODY
    ws4.cell(row=row, column=3, value=f'=IF({trade_rows_cell}=0,"-",B{row}/{trade_rows_cell})').font = BODY
    ws4.cell(row=row, column=3).number_format = '0.0%'
    row += 1
ws4.cell(row=row, column=1, value='Stop losses and take profits are almost never in the images — he writes them in the message text instead (323 messages mention an SL, 280 mention a TP/target).').font = Font(name=FONT, size=9, italic=True, color='808080')

row += 2
ws4.cell(row=row, column=1, value='Most traded symbols (by trade rows)').font = Font(name=FONT, bold=True, size=12)
row += 1
ws4.cell(row=row, column=1, value='Symbol').font = lab
ws4.cell(row=row, column=2, value='Rows').font = lab
ws4.cell(row=row, column=3, value='Long').font = lab
ws4.cell(row=row, column=4, value='Short').font = lab
ws4.cell(row=row, column=5, value='Avg ROI % shown').font = lab
row += 1
from collections import Counter
top = Counter(r[3] for r in rows if r[3]).most_common(15)
for base, _ in top:
    ws4.cell(row=row, column=1, value=base).font = BODY
    ws4.cell(row=row, column=2, value=f'=COUNTIF(Trades!$D$2:$D$100000,$A{row})').font = BODY
    ws4.cell(row=row, column=3, value=f'=COUNTIFS(Trades!$D$2:$D$100000,$A{row},Trades!$E$2:$E$100000,"long")').font = BODY
    ws4.cell(row=row, column=4, value=f'=COUNTIFS(Trades!$D$2:$D$100000,$A{row},Trades!$E$2:$E$100000,"short")').font = BODY
    c = ws4.cell(row=row, column=5, value=f'=IFERROR(AVERAGEIF(Trades!$D$2:$D$100000,$A{row},Trades!$K$2:$K$100000),"-")')
    c.font = BODY
    c.number_format = '#,##0.0'
    row += 1

row += 1
ws4.cell(row=row, column=1, value='Leverage used (where the image shows it)').font = Font(name=FONT, bold=True, size=12)
row += 1
ws4.cell(row=row, column=1, value='Leverage').font = lab
ws4.cell(row=row, column=2, value='Trade rows').font = lab
row += 1
levs = sorted({r[5] for r in rows if isinstance(r[5], (int, float))})
for lv in levs:
    ws4.cell(row=row, column=1, value=lv).font = BODY
    ws4.cell(row=row, column=2, value=f'=COUNTIF(Trades!$F$2:$F$100000,$A{row})').font = BODY
    row += 1

row += 1
ws4.cell(row=row, column=1, value='How to read this workbook').font = Font(name=FONT, bold=True, size=12)
row += 1
for line in [
    'Trades — one row per position visible in an image (335 rows). Blank cells mean the value was not shown in that image; nothing here is inferred or calculated.',
    'Images — one row per image (320), including the message he posted with it, a verbatim transcription of the text in the image, and what a chart was showing.',
    'Charts — the 24 price charts only, with timeframe, indicators, levels and drawings.',
    'Confidence "medium" and the Issues column flag crops where the symbol, platform or a price had to be read from context rather than from the image itself.',
    'Image file names match the files in the images folder; the Discord link opens the original message.',
]:
    ws4.cell(row=row, column=1, value=line).font = Font(name=FONT, size=10)
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws4.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws4.row_dimensions[row].height = 28
    row += 1

for col, w in zip('ABCDEF', [42, 14, 14, 12, 18, 14]):
    ws4.column_dimensions[col].width = w
ws4.sheet_view.showGridLines = False

wb.save(OUT)
print('saved', OUT, '| trade rows', n_trades, '| image rows', len(r2), '| charts', len(r3))
